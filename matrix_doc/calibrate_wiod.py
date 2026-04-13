#!/usr/bin/env python3
"""
calibrate_wiod.py — Calibration of dependency matrix A from WIOD 2016 NIOT data.

Method: Leontief direct requirements coefficients
  a_raw[i][j] = X[i][j] / GO[j]   (intermediate use / gross output)
  a_norm[i][j] = a_raw[i][j] / a_raw[j][j]   (normalised by self-consumption)

Sectors:
  Energy    = D35
  Water     = E36
  Transport = H49 + H50 + H51 + H52 + H53  (aggregated)

Countries: RUS, DEU, USA  (year 2014)

Output:
  matrix_doc/A_calibrated_wiod.json
  results/figures/heatmap_A_calibrated_wiod.png
"""
import json
import math
from pathlib import Path

import openpyxl

try:
    import numpy as np
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from scipy.stats import spearmanr
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False
    print("[warn] scipy/numpy not fully available — Spearman ρ computed manually")


NIOTS_DIR = Path("/Users/kuprich/Documents/diploma_repo/diploma/matrix_doc/sources/NIOTS")
OUT_JSON  = Path("/Users/kuprich/Documents/diploma_repo/diploma/matrix_doc/A_calibrated_wiod.json")
FIG_DIR   = Path("/Users/kuprich/Documents/diploma_repo/diploma/results/figures")
FIG_DIR.mkdir(parents=True, exist_ok=True)

COUNTRIES = ["RUS", "DEU", "USA"]
YEAR = 2014
SECTORS = ["Energy", "Water", "Transport"]  # model sectors

# WIOD codes for each model sector
SECTOR_CODES = {
    "Energy":    ["D35"],
    "Water":     ["E36"],
    "Transport": ["H49", "H50", "H51", "H52", "H53"],
}

# Expert matrix v1 for Spearman comparison (rows/cols: Energy, Water, Transport)
A_V1 = [
    [0.10, 0.40, 0.30],   # Energy row
    [0.05, 0.10, 0.05],   # Water row
    [0.20, 0.20, 0.15],   # Transport row
]


def load_niot(country: str):
    """Return dict: {(year, code, origin): row_values_dict}"""
    path = NIOTS_DIR / f"{country}_NIOT_nov16.xlsx"
    if not path.exists():
        print(f"  [warn] {path} not found — skipping {country}")
        return None, None
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["National IO-tables"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))  # row 1: Year, Code, Description, Origin, A01, ...
    next(rows)                  # row 2: descriptions — skip
    col_idx = {h: i for i, h in enumerate(headers) if h is not None}
    return ws, col_idx, headers, list(rows)   # return pre-consumed iterator as list


def extract_submatrix(country: str, year: int = YEAR):
    """
    Returns:
      X: 3×3 numpy array of intermediate flows (rows = from-sector, cols = to-sector)
      GO: list of 3 gross output values
      available: list of bools (True if sector has valid data)
    """
    path = NIOTS_DIR / f"{country}_NIOT_nov16.xlsx"
    if not path.exists():
        print(f"  [skip] {country} — file not found")
        return None, None, None

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["National IO-tables"]
    all_rows = list(ws.iter_rows(values_only=True))
    headers = list(all_rows[0])
    col_idx = {h: i for i, h in enumerate(headers) if h is not None}

    data_rows = all_rows[2:]  # skip header + description rows

    def get_cell(row, code):
        idx = col_idx.get(code)
        if idx is None:
            return 0.0
        v = row[idx]
        return float(v) if v is not None else 0.0

    # For each model sector, collect summed flows
    # X[i][j] = sum over source codes of sector i, sum over dest codes of sector j
    # (intermediate use of sector-j products by sector-i)

    # Identify Transport column codes present
    transport_codes_present = [c for c in SECTOR_CODES["Transport"] if c in col_idx]
    transport_row_codes = SECTOR_CODES["Transport"]

    # Build indexed lookup: (year, code, origin) → row
    lookup = {}
    for row in data_rows:
        if row[0] is None:
            continue
        try:
            yr = int(row[0])
        except (TypeError, ValueError):
            continue
        code   = row[1]
        origin = row[3]
        if code is None:
            continue
        lookup[(yr, code, origin)] = row

    # Helper: get row value for given year/code/origin, summing column codes
    def get_x(from_codes, to_codes, origin="Domestic"):
        """Sum of X[from_code, to_code] over all from/to combinations."""
        total = 0.0
        for fc in from_codes:
            row = lookup.get((year, fc, origin))
            if row is None:
                continue
            for tc in to_codes:
                total += get_cell(row, tc)
        return total

    def get_go(sector_codes):
        """Gross output for sector (GO row, Origin=TOT)."""
        total = 0.0
        for c in sector_codes:
            row = lookup.get((year, "GO", "TOT"))
            if row is None:
                # Some files use Origin=Gross_output or similar
                for origin_try in ["TOT", "Total", "Gross output"]:
                    row = lookup.get((year, "GO", origin_try))
                    if row is not None:
                        break
            if row is not None:
                total += get_cell(row, c)
        return total

    # Extract 3×3 X matrix and GO vector
    # Order: Energy=0, Water=1, Transport=2
    sectors_codes = [
        SECTOR_CODES["Energy"],
        SECTOR_CODES["Water"],
        SECTOR_CODES["Transport"],
    ]

    X = [[0.0] * 3 for _ in range(3)]
    GO = [0.0] * 3

    for i, from_codes in enumerate(sectors_codes):
        for j, to_codes in enumerate(sectors_codes):
            X[i][j] = get_x(from_codes, to_codes)

    for j, codes in enumerate(sectors_codes):
        GO[j] = get_go(codes)

    # Availability: sector is available if GO > 0
    available = [go > 0 for go in GO]

    return X, GO, available


def compute_a_raw(X, GO):
    """a_raw[i][j] = X[i][j] / GO[j]. NaN if GO[j]=0."""
    n = 3
    a_raw = [[None] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if GO[j] > 0:
                a_raw[i][j] = X[i][j] / GO[j]
            else:
                a_raw[i][j] = None  # unavailable
    return a_raw


def compute_a_norm(a_raw):
    """a_norm[i][j] = a_raw[i][j] / a_raw[j][j]. Diagonal = 1.0."""
    n = 3
    a_norm = [[None] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if a_raw[i][j] is None or a_raw[j][j] is None or a_raw[j][j] == 0:
                a_norm[i][j] = None
            else:
                a_norm[i][j] = a_raw[i][j] / a_raw[j][j]
    return a_norm


def average_matrices(matrices):
    """Average over list of 3×3 matrices, ignoring None values."""
    n = 3
    avg = [[0.0] * n for _ in range(n)]
    counts = [[0] * n for _ in range(n)]
    for mat in matrices:
        for i in range(n):
            for j in range(n):
                if mat[i][j] is not None:
                    avg[i][j] += mat[i][j]
                    counts[i][j] += 1
    for i in range(n):
        for j in range(n):
            if counts[i][j] > 0:
                avg[i][j] = avg[i][j] / counts[i][j]
            else:
                avg[i][j] = None
    return avg, counts


def rescale_to_max(mat, max_val=0.5, skip_diagonal=True):
    """Rescale matrix so that max off-diagonal element = max_val."""
    n = 3
    cur_max = 0.0
    for i in range(n):
        for j in range(n):
            if skip_diagonal and i == j:
                continue
            v = mat[i][j]
            if v is not None and v > cur_max:
                cur_max = v
    if cur_max == 0:
        return mat, 1.0
    scale = max_val / cur_max
    result = [[None] * n for _ in range(n)]
    for i in range(n):
        for j in range(n):
            if mat[i][j] is not None:
                result[i][j] = mat[i][j] * scale
    return result, scale


def spearman_rho(mat_a, mat_b):
    """Compute Spearman ρ between off-diagonal elements of two 3×3 matrices."""
    pairs = []
    for i in range(3):
        for j in range(3):
            if i == j:
                continue
            va = mat_a[i][j]
            vb = mat_b[i][j]
            if va is None or vb is None:
                continue
            pairs.append((va, vb))
    if len(pairs) < 2:
        return None
    xs = [p[0] for p in pairs]
    ys = [p[1] for p in pairs]

    if HAS_SCIPY:
        rho, pval = spearmanr(xs, ys)
        return float(rho), float(pval)

    # Manual Spearman
    def rank(lst):
        sorted_vals = sorted(enumerate(lst), key=lambda x: x[1])
        ranks = [0.0] * len(lst)
        for r, (orig_idx, _) in enumerate(sorted_vals):
            ranks[orig_idx] = r + 1.0
        return ranks

    rx = rank(xs)
    ry = rank(ys)
    n = len(rx)
    d2 = sum((rx[k] - ry[k]) ** 2 for k in range(n))
    rho = 1 - 6 * d2 / (n * (n**2 - 1))
    return float(rho), None


def fmt(v, decimals=4):
    if v is None:
        return "  N/A  "
    return f"{v:8.4f}"


def print_matrix(mat, label, sectors=SECTORS):
    print(f"\n{label}")
    print(f"  {'':12s}" + "".join(f"  {s:>10s}" for s in sectors))
    print(f"  {'-'*56}")
    for i, row_label in enumerate(sectors):
        vals = "".join(fmt(mat[i][j]) for j in range(3))
        print(f"  {row_label:12s}{vals}")


def make_heatmap(mat, title, filename, sectors=SECTORS):
    if not HAS_SCIPY:
        print(f"  [skip heatmap] matplotlib not available")
        return
    import numpy as np

    data = np.array([[mat[i][j] if mat[i][j] is not None else 0.0
                      for j in range(3)] for i in range(3)], dtype=float)

    fig, ax = plt.subplots(figsize=(5.5, 4.5))
    im = ax.imshow(data, cmap="Blues", vmin=0, vmax=max(1.5, float(np.nanmax(data))))
    plt.colorbar(im, ax=ax, label="a_norm value")
    ax.set_xticks(range(3)); ax.set_yticks(range(3))
    ax.set_xticklabels(sectors, fontsize=10)
    ax.set_yticklabels(sectors, fontsize=10)
    ax.set_xlabel("Source sector j", fontsize=10)
    ax.set_ylabel("Destination sector i", fontsize=10)
    ax.set_title(title, fontsize=11, fontweight="bold")

    for i in range(3):
        for j in range(3):
            v = mat[i][j]
            s = f"{v:.3f}" if v is not None else "N/A"
            color = "white" if (v is not None and v > 0.8) else "black"
            ax.text(j, i, s, ha="center", va="center", fontsize=11, color=color, fontweight="bold")

    plt.tight_layout()
    out = FIG_DIR / filename
    plt.savefig(out, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  [save] {out}")


def main():
    print("=" * 65)
    print("WIOD 2016 NIOT Calibration — A dependency matrix (3×3)")
    print(f"Year: {YEAR} | Sectors: Energy (D35), Water (E36), Transport (H49-H53)")
    print("=" * 65)

    country_results = {}

    for country in COUNTRIES:
        print(f"\n{'─'*50}")
        print(f"  Country: {country}")

        X, GO, available = extract_submatrix(country, YEAR)
        if X is None:
            print(f"  [skip] {country} — no data")
            continue

        print(f"  Gross Output:")
        for k, s in enumerate(SECTORS):
            status = "✓" if available[k] else "✗ (GO=0, skipped)"
            print(f"    {s:12s}: GO = {GO[k]:.2f}  {status}")

        a_raw  = compute_a_raw(X, GO)
        a_norm = compute_a_norm(a_raw)

        print_matrix(a_raw,  f"  a_raw  [{country}, {YEAR}]  a_raw[i][j] = X[i][j] / GO[j]")
        print_matrix(a_norm, f"  a_norm [{country}, {YEAR}]  a_norm[i][j] = a_raw[i][j] / a_raw[j][j]")

        country_results[country] = {
            "X": X,
            "GO": GO,
            "available": available,
            "a_raw": a_raw,
            "a_norm": a_norm,
        }

    if not country_results:
        print("[error] No country data loaded — aborting")
        return

    # Average a_raw across countries (simple Leontief coefficients)
    raw_matrices = [r["a_raw"] for r in country_results.values()]
    A_raw_avg, counts_raw = average_matrices(raw_matrices)

    # Average a_norm across countries (diagonal-normalised)
    norm_matrices = [r["a_norm"] for r in country_results.values()]
    A_norm_avg, counts = average_matrices(norm_matrices)

    # Rescaled version: a_norm rescaled to max off-diagonal = 0.5 (for model use)
    A_scaled, scale_factor = rescale_to_max(A_norm_avg, max_val=0.5, skip_diagonal=True)

    # Also rescaled a_raw to max = 0.5
    A_raw_scaled, raw_scale = rescale_to_max(A_raw_avg, max_val=0.5, skip_diagonal=True)

    print(f"\n{'='*65}")
    print(f"  Countries used: {list(country_results.keys())}")
    print(f"  Note on RUS: Water sector (E36) has GO=0 in WIOD → excluded from Water column/row averages")

    print_matrix(A_raw_avg,
                 f"\n  [1] A_raw_avg (mean Leontief coefficients a_raw[i][j] = X[i][j]/GO[j])\n"
                 f"      Values already in [0,1]; directly usable as model parameters")

    print_matrix(A_norm_avg,
                 f"\n  [2] A_norm_avg (mean normalised by self-consumption a_raw[i][j]/a_raw[j][j])\n"
                 f"      Diagonal = 1.0 by construction; large off-diagonal for small-GO sectors (Water)")

    print_matrix(A_scaled,
                 f"\n  [3] A_scaled (A_norm_avg rescaled so max off-diagonal = 0.5)\n"
                 f"      Scale factor: {scale_factor:.6f}  — recommended for model use")

    print_matrix(A_raw_scaled,
                 f"\n  [4] A_raw_scaled (A_raw_avg rescaled so max off-diagonal = 0.5)\n"
                 f"      Scale factor: {raw_scale:.6f}  — alternative: simpler Leontief base")

    print(f"\n  Country counts per cell (N contributing to a_norm average):")
    print(f"  {'':12s}" + "".join(f"  {s:>10s}" for s in SECTORS))
    for i, row_label in enumerate(SECTORS):
        vals = "".join(f"  {counts[i][j]:>10d}" for j in range(3))
        print(f"  {row_label:12s}{vals}")

    # Spearman vs A_v1 — for A_raw_avg (more comparable with A_v1 scale)
    print(f"\n{'─'*65}")
    print("  Spearman ρ (off-diagonal): A_raw_avg vs A_expert_v1")
    rho_raw = spearman_rho(A_raw_avg, A_V1)
    if rho_raw:
        rho, pval = rho_raw
        pval_str = f"  p-value = {pval:.4f}" if pval is not None else ""
        print(f"  ρ(A_raw_avg, A_v1) = {rho:.4f}{pval_str}")

    print("  Spearman ρ (off-diagonal): A_norm_avg vs A_expert_v1")
    rho_result = spearman_rho(A_norm_avg, A_V1)
    if rho_result:
        rho, pval = rho_result
        pval_str = f"  p-value = {pval:.4f}" if pval is not None else ""
        print(f"  ρ(A_norm_avg, A_v1) = {rho:.4f}{pval_str}")

    print("  Spearman ρ (off-diagonal): A_scaled vs A_expert_v1")
    rho_scaled = spearman_rho(A_scaled, A_V1)
    if rho_scaled:
        rho, pval = rho_scaled
        pval_str = f"  p-value = {pval:.4f}" if pval is not None else ""
        print(f"  ρ(A_scaled,   A_v1) = {rho:.4f}{pval_str}")

    print_matrix(A_V1, "\n  A_expert_v1 (reference, for Spearman comparison)")

    # Sign agreement (A_raw_avg vs A_v1, off-diagonal)
    agree = 0; total_offdiag = 0
    for i in range(3):
        for j in range(3):
            if i == j:
                continue
            va = A_raw_avg[i][j]
            vb = A_V1[i][j]
            if va is None or vb is None:
                continue
            total_offdiag += 1
            if (va > 0) == (vb > 0):
                agree += 1
    print(f"\n  Sign agreement (A_raw_avg vs A_v1): {agree}/{total_offdiag} off-diagonal")

    # Key comparison table: A_raw_avg vs A_v1 vs A_scaled
    print(f"\n{'─'*65}")
    print("  Off-diagonal element comparison:")
    print(f"  {'i→j':16s}  {'A_v1':>8}  {'A_raw_avg':>10}  {'A_scaled':>10}  {'Rank dir':>10}")
    for i, si in enumerate(SECTORS):
        for j, sj in enumerate(SECTORS):
            if i == j:
                continue
            v1  = A_V1[i][j]
            vr  = A_raw_avg[i][j]
            vs  = A_scaled[i][j]
            dir_ok = "same" if (vr is not None and ((vr > v1) == (A_V1[i][j] > 0.15))) else "—"
            vs_str = f"{vs:.4f}" if vs is not None else "N/A"
            vr_str = f"{vr:.4f}" if vr is not None else "N/A"
            print(f"  {si}←{sj:12s}  {v1:8.4f}  {vr_str:>10s}  {vs_str:>10s}")

    # Save JSON
    def clean_mat(m):
        return [[round(v, 6) if v is not None else None for v in row] for row in m]

    out_data = {
        "metadata": {
            "source": "WIOD 2016 NIOT",
            "year": YEAR,
            "method": "Leontief direct requirements, normalised by self-consumption, rescaled to max_offdiag=0.5",
            "sectors": SECTORS,
            "sector_codes": SECTOR_CODES,
            "countries_used": list(country_results.keys()),
            "note_RUS_water": "E36 GO=0 in RUS WIOD data — excluded from Water sector averages",
        },
        "A_calibrated": clean_mat(A_scaled),          # primary: a_norm rescaled
        "A_raw_avg": clean_mat(A_raw_avg),             # simple Leontief avg
        "A_norm_avg": clean_mat(A_norm_avg),           # diagonal-normalised (large values)
        "A_raw_scaled": clean_mat(A_raw_scaled),       # Leontief rescaled to max=0.5
        "scale_factor_norm": round(scale_factor, 6),
        "scale_factor_raw": round(raw_scale, 6),
        "per_country": {
            c: {
                "GO": [round(g, 2) for g in r["GO"]],
                "available_sectors": r["available"],
                "a_raw":  clean_mat(r["a_raw"]),
                "a_norm": clean_mat(r["a_norm"]),
            }
            for c, r in country_results.items()
        },
        "A_expert_v1": A_V1,
        "spearman_rho_raw_vs_v1":    round(rho_raw[0], 4)    if rho_raw    else None,
        "spearman_rho_norm_vs_v1":   round(rho_result[0], 4) if rho_result else None,
        "spearman_rho_scaled_vs_v1": round(rho_scaled[0], 4) if rho_scaled else None,
        "sign_agreement_raw_vs_v1": f"{agree}/{total_offdiag}",
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out_data, indent=2, ensure_ascii=False))
    print(f"\n  [save] {OUT_JSON}")

    # Heatmaps
    make_heatmap(A_scaled,
                 f"A_calibrated WIOD {YEAR} (a_norm, rescaled max=0.5)\n{'+'.join(country_results.keys())}",
                 "heatmap_A_calibrated_wiod.png")
    make_heatmap(A_raw_avg,
                 f"A_raw_avg WIOD {YEAR} (Leontief direct)\n{'+'.join(country_results.keys())}",
                 "heatmap_A_raw_avg_wiod.png")

    # Per-country heatmaps
    for c, r in country_results.items():
        make_heatmap(r["a_norm"],
                     f"a_norm {c} (WIOD {YEAR})",
                     f"heatmap_a_norm_{c}_{YEAR}.png")

    print(f"\n{'='*65}")
    print("  DONE")
    print(f"{'='*65}\n")

    return A_scaled, A_raw_avg, country_results


if __name__ == "__main__":
    main()
